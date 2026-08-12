from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.makro.schema_registry import (
    build_schema_registry,
    build_schema_registry_from_directory,
    parse_makro_loadsheet,
)


def _sample_loadsheet(path: Path, *, vertical: str = "air_purifier") -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = vertical
    index = workbook.create_sheet("Index")

    blue = PatternFill("solid", fgColor="5B9BD5")
    green = PatternFill("solid", fgColor="70AD47")

    sheet["A1"] = "Model Name"
    sheet["A1"].fill = blue
    sheet["A2"] = "Single Text"
    sheet["A3"] = "Air Purifier"
    sheet["A4"] = "Product model name without the brand"

    sheet["B1"] = "Air Flow Level"
    sheet["B1"].fill = blue
    sheet["B2"] = "Numeric"
    sheet["B3"] = "1"
    sheet["B4"] = "Air flow setting level"

    sheet["C1"] = "Colour"
    sheet["C1"].fill = green
    sheet["C2"] = "Single Text - click here to get allowed values"
    sheet["C2"].hyperlink = "#Index!A1"
    sheet["C3"] = "White"
    sheet["C4"] = "Choose the product colour"

    index["A1"] = "Colour"
    index["A2"] = "White"
    index["A3"] = "Black"

    sheet["D1"] = "Material"
    sheet["D1"].fill = green
    sheet["D2"] = "Multi Text - click here to get allowed values"
    sheet["D3"] = "Plastic::Metal"
    sheet["D4"] = "Choose one or more materials"
    index["B1"] = "Material"
    index["B2"] = "Plastic"
    index["B3"] = "Metal"
    validation = DataValidation(type="list", formula1="=Index!$B$2:$B$3")
    sheet.add_data_validation(validation)
    validation.add("D5:D1000")

    workbook.save(path)
    return path


def test_official_rows_colours_and_index_allowed_values_are_preserved(tmp_path: Path):
    path = _sample_loadsheet(tmp_path / "air_purifier.xlsx")
    schema = parse_makro_loadsheet(path, vertical_hint="air_purifier")

    assert schema.vertical == "air_purifier"
    assert len(schema.fields) == 4
    by_key = {field.attribute_key: field for field in schema.fields}

    assert by_key["model_name"].requirement == "required"
    assert by_key["model_name"].field_type == "text"
    assert by_key["model_name"].example == "Air Purifier"
    assert by_key["model_name"].description == "Product model name without the brand"

    assert by_key["air_flow_level"].requirement == "required"
    assert by_key["air_flow_level"].field_type == "number"

    assert by_key["colour"].requirement == "optional"
    assert by_key["colour"].field_type == "single_select"
    assert by_key["colour"].allowed_values == ["White", "Black"]
    assert by_key["colour"].format_hyperlink == "#Index!A1"

    assert by_key["material"].field_type == "multi_select"
    assert by_key["material"].allowed_values == ["Plastic", "Metal"]


def test_registry_merges_vertical_occurrences_without_losing_per_vertical_contract(tmp_path: Path):
    one = _sample_loadsheet(tmp_path / "air_purifier.xlsx", vertical="air_purifier")
    two = _sample_loadsheet(tmp_path / "fan.xlsx", vertical="fan")
    schemas = [
        parse_makro_loadsheet(one, vertical_hint="air_purifier"),
        parse_makro_loadsheet(two, vertical_hint="fan"),
    ]
    registry = build_schema_registry(schemas)

    assert registry["stats"]["vertical_count"] == 2
    assert registry["stats"]["field_occurrence_count"] == 8
    assert registry["stats"]["unique_attribute_count"] == 4
    assert set(registry["verticals"]) == {"air_purifier", "fan"}
    assert registry["field_catalog"]["air_flow_level"]["verticals"] == ["air_purifier", "fan"]
    assert registry["field_catalog"]["colour"]["allowed_values"] == ["White", "Black"]


def test_directory_registry_uses_vertical_subdirectory_as_hint(tmp_path: Path):
    downloads = tmp_path / "downloads"
    vertical_dir = downloads / "air_purifier"
    vertical_dir.mkdir(parents=True)
    _sample_loadsheet(vertical_dir / "template.xlsx")

    registry, failures = build_schema_registry_from_directory(downloads)

    assert failures == []
    assert registry["stats"]["vertical_count"] == 1
    assert "air_purifier" in registry["verticals"]
    json.dumps(registry, ensure_ascii=False)
