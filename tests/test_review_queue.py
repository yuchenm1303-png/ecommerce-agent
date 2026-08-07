from __future__ import annotations

from openpyxl import load_workbook

from app.answer_resolver import CONFLICT, MISSING, NEEDS_REVIEW, RESOLVED
from app.resolution_engine import ResolutionRecord
from app.review_queue import build_review_queue, summarize_review_queue, write_review_queue_xlsx


def record(status: str, label: str, *, eligible: bool, number: str):
    return ResolutionRecord(
        attribute_key=label.casefold().replace(" ", "_"),
        label=label,
        status=status,
        answer=None,
        answer_values=[],
        qualifier=None,
        confidence=0.0,
        source_type=None,
        source_reference=None,
        evidence=None,
        detail=f"reason:{status}",
        eligible_for_autofill=eligible,
        question_number=number,
        question_category="IMAGE",
    )


def test_review_queue_contains_only_blocked_records_and_prioritizes_conflicts():
    items = build_review_queue(
        [
            record(MISSING, "Missing", eligible=False, number="3"),
            record(RESOLVED, "Ready", eligible=True, number="1"),
            record(NEEDS_REVIEW, "Review", eligible=False, number="2"),
            record(CONFLICT, "Conflict", eligible=False, number="4"),
        ]
    )

    assert [item.status for item in items] == [CONFLICT, NEEDS_REVIEW, MISSING]
    assert summarize_review_queue(items) == {
        "total": 3,
        "conflict": 1,
        "needs_review": 1,
        "missing": 1,
    }


def test_review_queue_xlsx_is_human_readable(tmp_path):
    items = build_review_queue([record(NEEDS_REVIEW, "Screen Size", eligible=False, number="7")])
    output = write_review_queue_xlsx(items, tmp_path / "review.xlsx")

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        sheet = workbook["Review Queue"]
        assert sheet.cell(2, 1).value == "7"
        assert sheet.cell(2, 2).value == "Screen Size"
        assert sheet.cell(2, 3).value == NEEDS_REVIEW
        summary = workbook["Summary"]
        assert summary.cell(2, 1).value == "total"
        assert summary.cell(2, 2).value == 1
    finally:
        workbook.close()
