from __future__ import annotations

from openpyxl import Workbook

from app.qa_catalog import load_question_catalog
from app.resolver_inputs import ResolutionInputSpec, build_resolution_inputs


def _write_customer_qa(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(1, 1, "SKU: 237581229555")
    sheet.cell(2, 1, "选定变体: M8 双录 + 64GB 内存卡")
    sheet.cell(3, 1, "供应商: 1688 offer 850845635717")
    sheet.append(["编号", "问题", "问题说明", "答案"])
    sheet.append(["1", "Model Number", "产品型号", ""])
    workbook.save(path)


def test_qa_catalog_keeps_product_context_before_header(tmp_path):
    path = tmp_path / "qa.xlsx"
    _write_customer_qa(path)

    catalog = load_question_catalog(path)

    assert catalog.header_row == 4
    assert "SKU: 237581229555" in catalog.preamble_text
    assert "选定变体: M8 双录 + 64GB 内存卡" in catalog.preamble_text
    assert "1688 offer 850845635717" in catalog.preamble_text
    assert len(catalog.questions) == 1


def test_resolution_inputs_keep_preamble_as_grounded_customer_context(tmp_path):
    path = tmp_path / "qa.xlsx"
    _write_customer_qa(path)
    catalog = load_question_catalog(path)

    result = build_resolution_inputs(
        catalog,
        ResolutionInputSpec(sku="237581229555"),
    )

    assert "选定变体: M8 双录 + 64GB 内存卡" in result.bundle.supplemental_text
    assert any(item.startswith("customer_context_chars=") for item in result.warnings)

    sku_facts = result.bundle.candidates(("SKU",))
    assert any(
        fact.source_type == "customer_file" and fact.value == "237581229555"
        for fact in sku_facts
    )
