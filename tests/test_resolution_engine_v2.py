from __future__ import annotations

import json

from openpyxl import Workbook, load_workbook

from app.answer_resolver import CONFLICT, MISSING, NEEDS_REVIEW, RESOLVED
from app.evidence_pipeline import (
    add_fact,
    bundle_from_catalog_answers,
    bundle_from_facts_json,
)
from app.qa_catalog import load_question_catalog
from app.resolution_engine import ResolutionPolicy, resolve_catalog, resolve_one, summarize_resolution
from app.resolution_report import write_resolution_xlsx
from app.source_bundle import ProductSourceBundle


def _catalog_file(tmp_path):
    path = tmp_path / "vehicle_camera_system.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Vehicle Camera System Questions"])
    sheet.append(["Instructions"])
    sheet.append(["编号", "问题", "问题说明", "问题类别", "选项", "单位", "答案"])
    sheet.append([1, "Image Resolution", "camera output resolution", "IMAGE", "", "", ""])
    sheet.append([2, "White Balance", "", "IMAGE", "Auto; Daylight", "", ""])
    sheet.append([3, "Warranty Summary", "", "WARRANTY", "", "", "1 year"])
    workbook.save(path)
    return path


def test_catalog_keeps_blank_answer_rows_and_detects_row_3_header(tmp_path):
    catalog = load_question_catalog(_catalog_file(tmp_path))

    assert catalog.header_row == 3
    assert len(catalog.questions) == 3
    assert catalog.answered_count == 1
    assert catalog.unanswered_count == 2
    assert catalog.questions[1].options == ("Auto", "Daylight")
    assert catalog.questions[2].answer == "1 year"


def test_catalog_answers_become_explicit_evidence_only(tmp_path):
    catalog = load_question_catalog(_catalog_file(tmp_path))
    bundle = bundle_from_catalog_answers(catalog)

    assert len(bundle.evidence) == 1
    evidence = bundle.candidates(["Warranty Summary"])[0]
    assert evidence.value == "1 year"
    assert evidence.source_type == "customer_answer"


def test_resolution_report_preserves_all_questions(tmp_path):
    catalog = load_question_catalog(_catalog_file(tmp_path))
    bundle = bundle_from_catalog_answers(catalog)
    records = resolve_catalog(catalog, bundle)

    assert len(records) == 3
    assert [item.status for item in records] == [MISSING, MISSING, RESOLVED]
    summary = summarize_resolution(records)
    assert summary["total"] == 3
    assert summary["resolved"] == 1
    assert summary["missing"] == 2

    output = write_resolution_xlsx(records, tmp_path / "resolution.xlsx")
    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        sheet = workbook["Resolution"]
        assert sheet.max_row == 4
        assert sheet.cell(2, 2).value == "Image Resolution"
        assert sheet.cell(4, 7).value == "1 year"
    finally:
        workbook.close()


def test_low_confidence_is_blocked_even_when_value_resolves():
    bundle = ProductSourceBundle()
    add_fact(
        bundle,
        key="Image Resolution",
        value="1920x1080",
        source_type="ai_synthesis",
        source_reference="ai:claim=1",
        confidence=0.80,
    )
    field = {"attribute_key": "image_resolution", "label": "Image Resolution", "controls": []}

    record = resolve_one(field, bundle, policy=ResolutionPolicy(ai_auto_fill_min_confidence=0.92))

    assert record.status == NEEDS_REVIEW
    assert record.eligible_for_autofill is False
    assert record.answer == "1920x1080"
    assert "0.80" in record.detail


def test_conflicting_sources_are_blocked_with_provenance():
    bundle = ProductSourceBundle()
    add_fact(
        bundle,
        key="Screen Size",
        value="3.0 inch",
        source_type="supplier_doc",
        source_reference="supplier.pdf:p2",
    )
    add_fact(
        bundle,
        key="Screen Size",
        value="3.16 inch",
        source_type="product_image",
        source_reference="image:2",
    )
    field = {"attribute_key": "screen_size", "label": "Screen Size", "controls": []}

    record = resolve_one(field, bundle)

    assert record.status == CONFLICT
    assert record.eligible_for_autofill is False
    assert len(record.provenance) == 2


def test_facts_json_supports_aliases_for_manual_structured_facts(tmp_path):
    path = tmp_path / "facts.json"
    path.write_text(
        json.dumps(
            {
                "facts": [
                    {
                        "key": "Video Resolution",
                        "aliases": ["Image Resolution"],
                        "value": "1920x1080",
                        "source_type": "structured",
                        "source_reference": "manual-spec-table",
                        "confidence": 0.95,
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    bundle = bundle_from_facts_json(path)
    field = {"attribute_key": "image_resolution", "label": "Image Resolution", "controls": []}

    record = resolve_one(field, bundle)

    assert record.status == RESOLVED
    assert record.eligible_for_autofill is True
    assert record.answer == "1920x1080"
    assert record.source_reference == "manual-spec-table"
