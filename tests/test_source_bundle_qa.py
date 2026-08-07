from __future__ import annotations

from openpyxl import Workbook

from app.source_bundle import bundle_from_qa_file


def test_qa_excel_detects_chinese_headers_after_title_rows(tmp_path):
    path = tmp_path / "client-qa.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "问题表"
    sheet.append(["Vehicle Camera System 商品问题"])
    sheet.append(["以下内容用于上架"])
    sheet.append(["编号", "问题", "问题说明", "问题类别", "选项", "单位", "答案"])
    sheet.append([1, "Model Number", "", "", "", "", "L11"])
    sheet.append([2, "Ports", "", "", "", "", "USB-C"])
    workbook.save(path)

    bundle = bundle_from_qa_file(path)

    assert len(bundle.evidence) == 2
    model = bundle.candidates(["Model Number"])[0]
    assert model.value == "L11"
    assert "sheet=问题表" in model.source_reference
    assert "row=4" in model.source_reference


def test_qa_excel_searches_non_active_worksheet(tmp_path):
    path = tmp_path / "client-qa-multisheet.xlsx"
    workbook = Workbook()
    cover = workbook.active
    cover.title = "说明"
    cover.append(["客户资料说明"])
    data = workbook.create_sheet("QA")
    data.append(["编号", "问题", "答案"])
    data.append([1, "Recording Resolution", "1080p"])
    workbook.active = 0
    workbook.save(path)

    bundle = bundle_from_qa_file(path)

    answer = bundle.candidates(["Recording Resolution"])[0]
    assert answer.value == "1080p"
    assert "sheet=QA" in answer.source_reference
