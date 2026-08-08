from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .answer_resolver import CONFLICT, MISSING, NEEDS_REVIEW
from .resolution_engine import ResolutionRecord


_STATUS_PRIORITY = {
    CONFLICT: 10,
    NEEDS_REVIEW: 20,
    MISSING: 30,
}


@dataclass(slots=True)
class ReviewQueueItem:
    question_number: str
    label: str
    status: str
    answer: str
    confidence: float
    source_type: str
    source_reference: str
    evidence: str
    reason: str
    category: str
    unit: str
    options: list[str]
    provenance: list[dict[str, Any]]

    def as_dict(self) -> dict[str, Any]:
        return {
            "question_number": self.question_number,
            "label": self.label,
            "status": self.status,
            "answer": self.answer,
            "confidence": self.confidence,
            "source_type": self.source_type,
            "source_reference": self.source_reference,
            "evidence": self.evidence,
            "reason": self.reason,
            "category": self.category,
            "unit": self.unit,
            "options": self.options,
            "provenance": self.provenance,
        }


def build_review_queue(records: Iterable[ResolutionRecord]) -> list[ReviewQueueItem]:
    """Return every resolution record that is not eligible for automatic fill."""

    output: list[ReviewQueueItem] = []
    for record in records:
        if record.eligible_for_autofill:
            continue
        output.append(
            ReviewQueueItem(
                question_number=record.question_number,
                label=record.label,
                status=record.status,
                answer=record.answer or "",
                confidence=record.confidence,
                source_type=record.source_type or "",
                source_reference=record.source_reference or "",
                evidence=record.evidence or "",
                reason=record.detail,
                category=record.question_category,
                unit=record.question_unit,
                options=list(record.question_options),
                provenance=list(record.provenance),
            )
        )

    output.sort(
        key=lambda item: (
            _STATUS_PRIORITY.get(item.status, 99),
            item.category.casefold(),
            item.question_number,
            item.label.casefold(),
        )
    )
    return output


def summarize_review_queue(items: Iterable[ReviewQueueItem]) -> dict[str, int]:
    rows = list(items)
    return {
        "total": len(rows),
        "conflict": sum(item.status == CONFLICT for item in rows),
        "needs_review": sum(item.status == NEEDS_REVIEW for item in rows),
        "missing": sum(item.status == MISSING for item in rows),
    }


def write_review_queue_json(items: Iterable[ReviewQueueItem], path: str | Path) -> Path:
    rows = list(items)
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(
        json.dumps(
            {
                "summary": summarize_review_queue(rows),
                "items": [item.as_dict() for item in rows],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return target


def write_review_queue_xlsx(items: Iterable[ReviewQueueItem], path: str | Path) -> Path:
    rows = list(items)
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review Queue"
    headers = [
        "Question No.",
        "Question / Makro Label",
        "Status",
        "Candidate Answer",
        "Confidence",
        "Source Type",
        "Source Reference",
        "Evidence",
        "Reason / Required Action",
        "Category",
        "Unit",
        "Allowed Options",
        "Provenance JSON",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    for item in rows:
        sheet.append(
            [
                item.question_number,
                item.label,
                item.status,
                item.answer,
                round(item.confidence, 4),
                item.source_type,
                item.source_reference,
                item.evidence,
                item.reason,
                item.category,
                item.unit,
                " | ".join(item.options),
                json.dumps(item.provenance, ensure_ascii=False),
            ]
        )

    if rows:
        sheet.auto_filter.ref = f"A1:M{len(rows) + 1}"
    widths = [12, 34, 18, 30, 12, 20, 54, 42, 70, 22, 12, 45, 80]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    summary = workbook.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    for key, value in summarize_review_queue(rows).items():
        summary.append([key, value])
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 14

    workbook.save(target)
    return target
