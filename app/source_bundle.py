from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .data_loader import load_products


QUESTION_HEADERS = {
    "question",
    "questions",
    "attribute",
    "attribute name",
    "field",
    "field name",
    "property",
    "问题",
    "属性",
    "属性名",
    "字段",
    "字段名",
}
ANSWER_HEADERS = {
    "answer",
    "answers",
    "value",
    "attribute value",
    "答案",
    "回答",
    "值",
    "属性值",
}


def normalize_key(value: object) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", text.casefold())


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@dataclass(slots=True, frozen=True)
class SourceEvidence:
    """One explicit piece of product evidence.

    ``evidence_text`` is the source-grounded snippet/visual transcription that
    justifies the value. ``note`` is reserved for pipeline/audit metadata such as
    confidence capping. Keeping them separate prevents reports from confusing a
    resolver value with the actual supporting evidence.
    """

    key: str
    value: str | tuple[str, ...]
    source_type: str
    source_reference: str
    priority: int
    confidence: float = 1.0
    evidence_text: str = ""
    note: str = ""

    @property
    def normalized_key(self) -> str:
        return normalize_key(self.key)


@dataclass(slots=True)
class ProductSourceBundle:
    sku: str = ""
    evidence: list[SourceEvidence] = field(default_factory=list)
    image_paths: tuple[str, ...] = ()
    product_url: str | None = None
    supplemental_text: str = ""

    def add_evidence(
        self,
        *,
        key: str,
        value: str | Iterable[str],
        source_type: str,
        source_reference: str,
        priority: int,
        confidence: float = 1.0,
        evidence_text: str = "",
        note: str = "",
    ) -> None:
        if isinstance(value, str):
            stored: str | tuple[str, ...] = value.strip()
            if not stored:
                return
        else:
            stored = tuple(str(item).strip() for item in value if str(item).strip())
            if not stored:
                return
        self.evidence.append(
            SourceEvidence(
                key=key.strip(),
                value=stored,
                source_type=source_type,
                source_reference=source_reference,
                priority=priority,
                confidence=confidence,
                evidence_text=evidence_text.strip(),
                note=note,
            )
        )

    def candidates(self, keys: Iterable[str]) -> list[SourceEvidence]:
        wanted = {normalize_key(key) for key in keys if key}
        if not wanted:
            return []
        return [item for item in self.evidence if item.normalized_key in wanted]


def bundle_from_product_table(
    path: str | Path,
    *,
    sku: str | None = None,
    image_paths: Iterable[str] = (),
    product_url: str | None = None,
    supplemental_text: str = "",
) -> ProductSourceBundle:
    """Load one product row from the existing CSV/XLSX product-table format."""

    source = Path(path)
    records = load_products(source)
    if sku:
        matches = [record for record in records if record.sku == sku]
        if not matches:
            raise ValueError(f"数据文件中未找到 SKU={sku}")
        record = matches[0]
    elif len(records) == 1:
        record = records[0]
    else:
        raise ValueError("数据文件包含多个商品，请通过 --sku 指定一个 SKU。")

    bundle = ProductSourceBundle(
        sku=record.sku,
        image_paths=tuple(str(item) for item in image_paths),
        product_url=product_url,
        supplemental_text=supplemental_text,
    )
    bundle.add_evidence(
        key="SKU",
        value=record.sku,
        source_type="structured",
        source_reference=f"{source.name}:row={record.row_number}",
        priority=10,
        evidence_text=f"SKU={record.sku}",
    )
    for key, value in record.values.items():
        bundle.add_evidence(
            key=key,
            value=value,
            source_type="structured",
            source_reference=f"{source.name}:row={record.row_number}:column={key}",
            priority=10,
            evidence_text=f"{key}={value}",
        )
    return bundle


def _find_header(headers: list[str], aliases: set[str]) -> int | None:
    normalized_aliases = {normalize_key(item) for item in aliases}
    for index, header in enumerate(headers):
        if normalize_key(header) in normalized_aliases:
            return index
    return None


def _looks_like_qa_header(row: Iterable[Any]) -> bool:
    headers = [_stringify(item) for item in row]
    return (
        _find_header(headers, QUESTION_HEADERS) is not None
        and _find_header(headers, ANSWER_HEADERS) is not None
    )


def _locate_qa_table(
    rows: list[list[Any]], *, max_header_rows: int = 50
) -> tuple[list[str], list[list[Any]], int]:
    """Find the actual QA header row instead of assuming row 1."""

    if not rows:
        raise ValueError("QA 文件为空。")
    for index, row in enumerate(rows[:max_header_rows]):
        if _looks_like_qa_header(row):
            return [_stringify(item) for item in row], rows[index + 1 :], index + 1
    raise ValueError(
        "未识别到 QA 列。请确认前 50 行内存在同时包含 Question/Attribute/问题/字段 "
        "与 Answer/Value/答案 的表头行。"
    )


def _qa_rows_from_csv(path: Path) -> tuple[list[str], list[list[Any]], int, str | None]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = [list(row) for row in csv.reader(handle)]
    headers, data_rows, header_row = _locate_qa_table(rows)
    return headers, data_rows, header_row, None


def _qa_rows_from_excel(path: Path) -> tuple[list[str], list[list[Any]], int, str | None]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        saw_any_rows = False
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if rows:
                saw_any_rows = True
            try:
                headers, data_rows, header_row = _locate_qa_table(rows)
            except ValueError:
                continue
            return headers, data_rows, header_row, sheet.title
        if not saw_any_rows:
            raise ValueError("QA Excel 文件为空。")
        raise ValueError(
            "未识别到 QA 列。已检查所有工作表前 50 行；请确认存在同时包含问题/字段与答案的表头。"
        )
    finally:
        workbook.close()


def bundle_from_qa_file(
    path: str | Path,
    *,
    sku: str = "",
    image_paths: Iterable[str] = (),
    product_url: str | None = None,
    supplemental_text: str = "",
) -> ProductSourceBundle:
    """Load a question/answer workbook like the client's current manual workflow."""

    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        headers, rows, header_row, sheet_name = _qa_rows_from_csv(source)
    elif suffix in {".xlsx", ".xlsm"}:
        headers, rows, header_row, sheet_name = _qa_rows_from_excel(source)
    else:
        raise ValueError("QA 文件当前仅支持 .csv / .xlsx / .xlsm。")

    question_index = _find_header(headers, QUESTION_HEADERS)
    answer_index = _find_header(headers, ANSWER_HEADERS)
    if question_index is None or answer_index is None:
        raise ValueError(
            "未识别到 QA 列。请至少包含 Question/Attribute/字段 与 Answer/Value/答案 列。"
        )

    bundle = ProductSourceBundle(
        sku=sku,
        image_paths=tuple(str(item) for item in image_paths),
        product_url=product_url,
        supplemental_text=supplemental_text,
    )
    for row_number, row in enumerate(rows, start=header_row + 1):
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        question = _stringify(padded[question_index])
        answer = _stringify(padded[answer_index])
        if not question or not answer:
            continue
        location = f"{source.name}:row={row_number}"
        if sheet_name:
            location = f"{source.name}:sheet={sheet_name}:row={row_number}"
        bundle.add_evidence(
            key=question,
            value=answer,
            source_type="customer_file",
            source_reference=location,
            priority=20,
            evidence_text=f"{question}={answer}",
        )
    if not bundle.evidence:
        raise ValueError("QA 文件中没有可用的明确答案。")
    return bundle
