from __future__ import annotations

import json
import re
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter, range_boundaries


_INDEX_SHEET_HINTS = ("index", "allowed", "lookup", "values")
_NON_VERTICAL_SHEET_HINTS = ("index", "instruction", "read me", "readme", "cover", "guide")
_PLACEHOLDER_OPTIONS = {
    "select",
    "select one",
    "select vertical",
    "choose",
    "choose one",
    "please select",
}


class MakroSchemaParseError(ValueError):
    """Raised when a Makro loadsheet cannot be interpreted safely."""


def _text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _text(value).casefold()).strip("_")


def _rgb_from_color(color: Any) -> str:
    colour_type = str(getattr(color, "type", "") or "")
    if colour_type == "rgb":
        raw = str(getattr(color, "rgb", "") or "")
        return raw[-6:].upper() if len(raw) >= 6 else ""
    if colour_type == "indexed":
        try:
            indexed = int(getattr(color, "indexed"))
            raw = COLOR_INDEX[indexed]
        except (TypeError, ValueError, IndexError):
            return ""
        return str(raw)[-6:].upper()
    return ""


def _fill_token(cell: Any) -> str:
    fill = getattr(cell, "fill", None)
    if fill is None:
        return ""
    rgb = _rgb_from_color(getattr(fill, "fgColor", None))
    if rgb:
        return f"rgb:{rgb}"
    colour = getattr(fill, "fgColor", None)
    colour_type = str(getattr(colour, "type", "") or "")
    if colour_type == "theme":
        return f"theme:{getattr(colour, 'theme', '')}:tint={getattr(colour, 'tint', '')}"
    if colour_type == "indexed":
        return f"indexed:{getattr(colour, 'indexed', '')}"
    return colour_type


def _requirement_from_header(cell: Any) -> str:
    """Map official blue/green header semantics without guessing unknown colours.

    Makro's help centre documents blue headers as mandatory and green headers as
    optional. RGB/indexed fills can be classified mechanically by colour family;
    theme colours that cannot be resolved safely remain ``unknown``.
    """

    rgb = _rgb_from_color(getattr(getattr(cell, "fill", None), "fgColor", None))
    if not rgb or len(rgb) != 6:
        return "unknown"
    try:
        red, green, blue = (int(rgb[index : index + 2], 16) for index in (0, 2, 4))
    except ValueError:
        return "unknown"
    if blue >= red + 25 and blue >= green + 10:
        return "required"
    if green >= red + 25 and green >= blue + 10:
        return "optional"
    return "unknown"


def _is_index_sheet(title: str) -> bool:
    normalized = _text(title).casefold()
    return any(token in normalized for token in _INDEX_SHEET_HINTS)


def _is_non_vertical_sheet(title: str) -> bool:
    normalized = _text(title).casefold()
    return any(token in normalized for token in _NON_VERTICAL_SHEET_HINTS)


def _sheet_score(sheet: Any) -> int:
    return sum(1 for cell in sheet[1] if _text(cell.value))


def _pick_vertical_sheet(workbook: Any, vertical_hint: str = "") -> Any:
    wanted = _key(vertical_hint)
    candidates = [sheet for sheet in workbook.worksheets if not _is_non_vertical_sheet(sheet.title)]
    if wanted:
        exact = [sheet for sheet in candidates if _key(sheet.title) == wanted]
        if len(exact) == 1:
            return exact[0]
    ranked = sorted(candidates, key=lambda sheet: (_sheet_score(sheet), sheet.max_column), reverse=True)
    if not ranked or _sheet_score(ranked[0]) == 0:
        raise MakroSchemaParseError("loadsheet 中找不到包含 row 1 字段标题的 Vertical 工作表。")
    return ranked[0]


def _parse_sheet_ref(reference: str) -> tuple[str, str] | None:
    value = _text(reference).lstrip("#")
    match = re.match(r"^(?:'([^']+)'|([^!]+))!(.+)$", value)
    if not match:
        return None
    sheet_name = _text(match.group(1) or match.group(2))
    cell_ref = _text(match.group(3)).replace("$", "")
    return (sheet_name, cell_ref) if sheet_name and cell_ref else None


def _range_values(workbook: Any, reference: str, *, header_label: str = "") -> list[str]:
    parsed = _parse_sheet_ref(reference)
    if parsed is None:
        return []
    sheet_name, cell_ref = parsed
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]

    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_ref)
    except ValueError:
        return []

    # Hyperlinks often point to the allowed-values column header rather than an
    # explicit finite range. Walk down that one column until the first sustained
    # blank region. Explicit ranges keep their declared end row.
    explicit_range = ":" in cell_ref
    if not explicit_range:
        max_col = min_col
        max_row = sheet.max_row

    output: list[str] = []
    blank_run = 0
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            value = _text(sheet.cell(row=row, column=col).value)
            if not value:
                blank_run += 1
                continue
            blank_run = 0
            if row == min_row and _key(value) in {_key(header_label), "allowed_values", "values", "value"}:
                continue
            if value.casefold() not in _PLACEHOLDER_OPTIONS and value not in output:
                output.append(value)
        if not explicit_range and output and blank_run >= 3:
            break
    return output


def _hyperlink_values(workbook: Any, cell: Any, *, header_label: str) -> tuple[list[str], str]:
    hyperlink = getattr(cell, "hyperlink", None)
    if hyperlink is None:
        return [], ""
    reference = _text(getattr(hyperlink, "location", "")) or _text(getattr(hyperlink, "target", ""))
    if not reference or reference.startswith(("http://", "https://")):
        return [], reference
    return _range_values(workbook, reference, header_label=header_label), reference


def _validation_values(workbook: Any, sheet: Any, column: int, *, header_label: str) -> list[str]:
    coordinate = f"{get_column_letter(column)}5"
    validations = getattr(getattr(sheet, "data_validations", None), "dataValidation", []) or []
    for validation in validations:
        try:
            if coordinate not in validation.cells:
                continue
        except TypeError:
            continue
        formula = _text(getattr(validation, "formula1", ""))
        if not formula:
            continue
        if formula.startswith('"') and formula.endswith('"'):
            return [
                item.strip()
                for item in formula[1:-1].split(",")
                if item.strip() and item.strip().casefold() not in _PLACEHOLDER_OPTIONS
            ]
        formula = formula.lstrip("=")
        values = _range_values(workbook, formula, header_label=header_label)
        if values:
            return values
    return []


def _index_columns(workbook: Any) -> dict[str, list[str]]:
    domains: dict[str, list[str]] = {}
    for sheet in workbook.worksheets:
        if not _is_index_sheet(sheet.title):
            continue
        for column in range(1, sheet.max_column + 1):
            header_row = 0
            header = ""
            for row in range(1, min(sheet.max_row, 10) + 1):
                candidate = _text(sheet.cell(row=row, column=column).value)
                if candidate:
                    header_row = row
                    header = candidate
                    break
            if not header:
                continue
            values: list[str] = []
            for row in range(header_row + 1, sheet.max_row + 1):
                value = _text(sheet.cell(row=row, column=column).value)
                if value and value.casefold() not in _PLACEHOLDER_OPTIONS and value not in values:
                    values.append(value)
            if values:
                domains[_key(header)] = values
    return domains


def _format_kind(format_text: str, allowed_values: Iterable[str], label: str) -> str:
    normalized = _text(format_text).casefold()
    allowed = list(allowed_values)
    if allowed:
        return "multi_select" if "multi" in normalized else "single_select"
    if "multi text" in normalized or "multi-text" in normalized:
        return "multi_text"
    if any(token in normalized for token in ("number", "numeric", "decimal", "integer", "float")):
        return "number"
    if any(token in normalized for token in ("yes/no", "boolean", "true/false")):
        return "boolean"
    if "date" in normalized:
        return "date"
    if "url" in normalized or "link" in normalized:
        return "image_url" if "image" in _text(label).casefold() or "image" in normalized else "url"
    if "text" in normalized or "string" in normalized:
        return "text"
    return "unknown"


@dataclass(slots=True)
class MakroLoadsheetField:
    attribute_key: str
    label: str
    format_text: str
    field_type: str
    example: str
    description: str
    requirement: str
    allowed_values: list[str] = field(default_factory=list)
    format_hyperlink: str = ""
    header_fill: str = ""
    column_index: int = 0

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(slots=True)
class MakroVerticalSchema:
    vertical: str
    sheet_name: str
    source_file: str
    fields: list[MakroLoadsheetField]

    def stats(self) -> dict[str, Any]:
        requirements = Counter(field.requirement for field in self.fields)
        types = Counter(field.field_type for field in self.fields)
        return {
            "field_count": len(self.fields),
            "required": requirements.get("required", 0),
            "optional": requirements.get("optional", 0),
            "unknown_requirement": requirements.get("unknown", 0),
            "field_types": dict(sorted(types.items())),
        }

    def as_dict(self) -> dict[str, Any]:
        return {
            "vertical": self.vertical,
            "sheet_name": self.sheet_name,
            "source_file": self.source_file,
            "stats": self.stats(),
            "fields": [field.as_dict() for field in self.fields],
        }


def parse_makro_loadsheet(path: str | Path, *, vertical_hint: str = "") -> MakroVerticalSchema:
    source = Path(path)
    if source.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise MakroSchemaParseError(f"暂不支持 {source.suffix or '无扩展名'}；Makro schema parser 只读取 xlsx/xlsm。")
    workbook = load_workbook(source, data_only=False, read_only=False)
    sheet = _pick_vertical_sheet(workbook, vertical_hint=vertical_hint)
    index_domains = _index_columns(workbook)

    fields: list[MakroLoadsheetField] = []
    seen_keys: Counter[str] = Counter()
    for column in range(1, sheet.max_column + 1):
        header_cell = sheet.cell(row=1, column=column)
        label = _text(header_cell.value)
        if not label:
            continue
        format_cell = sheet.cell(row=2, column=column)
        format_text = _text(format_cell.value)
        example = _text(sheet.cell(row=3, column=column).value)
        description = _text(sheet.cell(row=4, column=column).value)

        hyperlink_values, hyperlink_ref = _hyperlink_values(workbook, format_cell, header_label=label)
        allowed_values = hyperlink_values or _validation_values(
            workbook, sheet, column, header_label=label
        )
        if not allowed_values:
            allowed_values = list(index_domains.get(_key(label), []))

        base_key = _key(label) or f"column_{column}"
        seen_keys[base_key] += 1
        attribute_key = base_key if seen_keys[base_key] == 1 else f"{base_key}__{seen_keys[base_key]}"
        fields.append(
            MakroLoadsheetField(
                attribute_key=attribute_key,
                label=label,
                format_text=format_text,
                field_type=_format_kind(format_text, allowed_values, label),
                example=example,
                description=description,
                requirement=_requirement_from_header(header_cell),
                allowed_values=allowed_values,
                format_hyperlink=hyperlink_ref,
                header_fill=_fill_token(header_cell),
                column_index=column,
            )
        )

    if not fields:
        raise MakroSchemaParseError(f"{source.name} 的 {sheet.title!r} 工作表没有可解析字段。")
    vertical = _text(vertical_hint) or sheet.title
    return MakroVerticalSchema(
        vertical=vertical,
        sheet_name=sheet.title,
        source_file=str(source.resolve()),
        fields=fields,
    )


def _catalog(verticals: Iterable[MakroVerticalSchema]) -> dict[str, dict[str, Any]]:
    output: dict[str, dict[str, Any]] = {}
    for vertical in verticals:
        for item in vertical.fields:
            key = _key(item.label) or item.attribute_key
            entry = output.setdefault(
                key,
                {
                    "labels": [],
                    "verticals": [],
                    "field_types": [],
                    "format_texts": [],
                    "required_in": [],
                    "optional_in": [],
                    "unknown_requirement_in": [],
                    "allowed_values": [],
                    "examples": [],
                    "descriptions": [],
                },
            )
            for bucket, value in (
                ("labels", item.label),
                ("verticals", vertical.vertical),
                ("field_types", item.field_type),
                ("format_texts", item.format_text),
                ("examples", item.example),
                ("descriptions", item.description),
            ):
                if value and value not in entry[bucket]:
                    entry[bucket].append(value)
            requirement_bucket = {
                "required": "required_in",
                "optional": "optional_in",
            }.get(item.requirement, "unknown_requirement_in")
            if vertical.vertical not in entry[requirement_bucket]:
                entry[requirement_bucket].append(vertical.vertical)
            for value in item.allowed_values:
                if value not in entry["allowed_values"]:
                    entry["allowed_values"].append(value)
    return dict(sorted(output.items()))


def build_schema_registry(verticals: Iterable[MakroVerticalSchema]) -> dict[str, Any]:
    schemas = list(verticals)
    type_counts = Counter(field.field_type for schema in schemas for field in schema.fields)
    requirement_counts = Counter(field.requirement for schema in schemas for field in schema.fields)
    catalog = _catalog(schemas)
    return {
        "contract_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "Makro Marketplace vertical-specific Bulk Product Creation loadsheets",
        "stats": {
            "vertical_count": len(schemas),
            "field_occurrence_count": sum(len(schema.fields) for schema in schemas),
            "unique_attribute_count": len(catalog),
            "field_types": dict(sorted(type_counts.items())),
            "requirements": dict(sorted(requirement_counts.items())),
        },
        "verticals": {schema.vertical: schema.as_dict() for schema in sorted(schemas, key=lambda s: s.vertical.casefold())},
        "field_catalog": catalog,
    }


def build_schema_registry_from_directory(root: str | Path) -> tuple[dict[str, Any], list[dict[str, str]]]:
    directory = Path(root)
    schemas: list[MakroVerticalSchema] = []
    failures: list[dict[str, str]] = []
    files = sorted(
        [*directory.rglob("*.xlsx"), *directory.rglob("*.xlsm")],
        key=lambda path: str(path).casefold(),
    )
    for path in files:
        vertical_hint = path.parent.name if path.parent != directory else ""
        try:
            schemas.append(parse_makro_loadsheet(path, vertical_hint=vertical_hint))
        except Exception as exc:
            failures.append({"file": str(path.resolve()), "error": str(exc)})
    if not schemas:
        raise MakroSchemaParseError(f"{directory} 中没有成功解析任何 Makro xlsx/xlsm loadsheet。")
    return build_schema_registry(schemas), failures


def write_schema_registry(payload: dict[str, Any], path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return target
