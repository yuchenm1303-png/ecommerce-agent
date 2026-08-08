from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .fill_plan import LiveFillPlan


PLAN_COLUMNS = (
    "Section",
    "Attribute Key",
    "Makro Label",
    "Required",
    "Action",
    "Answer",
    "Resolution Status",
    "Auto Fill Eligible",
    "Review Preview Eligible",
    "Gate Reason",
    "Confidence",
    "Source Type",
    "Source Reference",
    "Evidence",
    "Reason",
    "Provenance JSON",
)


def write_fill_plan_json(plan: LiveFillPlan, path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(
        json.dumps(plan.as_dict(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return target


def write_fill_plan_xlsx(plan: LiveFillPlan, path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fill Plan"
    sheet.append(list(PLAN_COLUMNS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    for item in plan.items:
        resolution = item.resolution
        sheet.append(
            [
                item.section_heading,
                item.attribute_key,
                item.label,
                "YES" if item.required else "NO",
                item.action,
                resolution.answer or "",
                resolution.status,
                "YES" if resolution.eligible_for_autofill else "NO",
                "YES" if resolution.preview_eligible else "NO",
                resolution.gate_reason,
                round(resolution.confidence, 4),
                resolution.source_type or "",
                resolution.source_reference or "",
                resolution.evidence or "",
                item.reason,
                json.dumps(resolution.provenance, ensure_ascii=False),
            ]
        )

    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(PLAN_COLUMNS))}{max(1, len(plan.items) + 1)}"
    )
    widths = {
        1: 34,
        2: 30,
        3: 32,
        4: 10,
        5: 12,
        6: 30,
        7: 18,
        8: 16,
        9: 22,
        10: 28,
        11: 12,
        12: 20,
        13: 54,
        14: 48,
        15: 58,
        16: 72,
    }
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width

    summary = workbook.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    for key, value in plan.summary().items():
        summary.append(
            [key, json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value]
        )
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 48

    if plan.warnings:
        warnings = workbook.create_sheet("Warnings")
        warnings.append(["Warning"])
        warnings["A1"].font = Font(bold=True)
        for warning in plan.warnings:
            warnings.append([warning])
        warnings.column_dimensions["A"].width = 100

    workbook.save(target)
    return target
