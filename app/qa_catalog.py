from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .source_bundle import ANSWER_HEADERS, QUESTION_HEADERS, normalize_key


NUMBER_HEADERS = {"编号", "序号", "number", "no", "no.", "id"}
EXPLANATION_HEADERS = {
    "问题说明",
    "说明",
    "解释",
    "description",
    "explanation",
    "help",
    "hint",
}
CATEGORY_HEADERS = {"问题类别", "类别", "分类", "category", "type"}
OPTION_HEADERS = {"选项", "可选项", "options", "option", "allowed values", "allowed_values"}
UNIT_HEADERS = {"单位", "unit", "units", "qualifier"}


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _find_header(headers: list[str], aliases: Iterable[str]) -> int | None:
    wanted = {normalize_key(alias) for alias in aliases}
    for index, header in enumerate(headers):
        if normalize_key(header) in wanted:
            return index
    return None


def _split_options(value: str) -> tuple[str, ...]:
    text = value.strip()
    if not text:
        return ()
    parts = [item.strip() for item in re.split(r"[\n|;；]+", text) if item.strip()]
    if len(parts) <= 1:
        return (text,)
    return tuple(dict.fromkeys(parts))


def _preamble_text(rows: list[list[Any]], header_row: int) -> str:
    """Keep non-empty customer context before the QA header.

    Customer workbooks often place the SKU, exact selected variant, supplier URL
    and copy/compliance instructions above the question table. Dropping those
    rows silently removes some of the strongest product-identity evidence before
    semantic extraction even starts.
    """

    lines: list[str] = []
    for row in rows[: max(0, header_row - 1)]:
        values = [_stringify(value) for value in row]
        values = [value for value in values if value]
        if values:
            lines.append(" | ".join(values))
    return "\n".join(lines).strip()


@dataclass(slots=True, frozen=True)
class QuestionRecord:
    number: str
    question: str
    explanation: str = ""
    category: str = ""
    options: tuple[str, ...] = ()
    unit: str = ""
    answer: str = ""
    source_reference: str = ""
    row_number: int = 0
    extra: dict[str, str] = field(default_factory=dict)

    @property
    def normalized_question(self) -> str:
        return normalize_key(self.question)

    @property
    def has_answer(self) -> bool:
        return bool(self.answer.strip())

    def as_semantic_field(self) -> dict[str, Any]:
        options = [{"text": item, "value": item} for item in self.options]
        controls: list[dict[str, Any]] = []
        if options:
            controls.append(
                {
                    "id": "",
                    "name": "",
                    "field_kind": "select",
                    "options": options,
                }
            )
        return {
            "attribute_key": self.question,
            "label": self.question,
            "section_heading": self.category,
            "required": False,
            "multi_value": False,
            "options": options,
            "controls": controls,
            "question_explanation": self.explanation,
            "question_unit": self.unit,
            "question_number": self.number,
        }


@dataclass(slots=True)
class QuestionCatalog:
    source_path: str
    sheet_name: str | None
    header_row: int
    questions: list[QuestionRecord]
    preamble_text: str = ""

    @property
    def answered_count(self) -> int:
        return sum(1 for item in self.questions if item.has_answer)

    @property
    def unanswered_count(self) -> int:
        return len(self.questions) - self.answered_count

    def by_question(self, question: str) -> QuestionRecord | None:
        wanted = normalize_key(question)
        for item in self.questions:
            if item.normalized_question == wanted:
                return item
        return None


def _looks_like_header(row: Iterable[Any]) -> bool:
    headers = [_stringify(item) for item in row]
    if _find_header(headers, QUESTION_HEADERS) is None:
        return False
    companion_groups = (
        ANSWER_HEADERS,
        NUMBER_HEADERS,
        EXPLANATION_HEADERS,
        CATEGORY_HEADERS,
        OPTION_HEADERS,
        UNIT_HEADERS,
    )
    return any(_find_header(headers, aliases) is not None for aliases in companion_groups)


def _locate_table(rows: list[list[Any]], max_header_rows: int = 50) -> tuple[list[str], list[list[Any]], int]:
    if not rows:
        raise ValueError("QA 文件为空。")
    for index, row in enumerate(rows[:max_header_rows]):
        if _looks_like_header(row):
            return [_stringify(item) for item in row], rows[index + 1 :], index + 1
    raise ValueError(
        "前 50 行内未识别到完整 QA 表头。表头必须包含 Question/问题/字段，"
        "以及 Answer/编号/说明/类别/选项/单位 中至少一列。"
    )


def _load_rows(
    path: Path,
) -> tuple[list[str], list[list[Any]], int, str | None, str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [list(row) for row in csv.reader(handle)]
        headers, data, header_row = _locate_table(rows)
        return headers, data, header_row, None, _preamble_text(rows, header_row)

    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("QA question catalog 仅支持 .csv / .xlsx / .xlsm。")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            try:
                headers, data, header_row = _locate_table(rows)
            except ValueError:
                continue
            return (
                headers,
                data,
                header_row,
                sheet.title,
                _preamble_text(rows, header_row),
            )
    finally:
        workbook.close()
    raise ValueError("已检查所有工作表，未识别到 QA 问题表。")


def load_question_catalog(path: str | Path) -> QuestionCatalog:
    """Load every question row plus customer context above the table."""

    source = Path(path)
    headers, rows, header_row, sheet_name, preamble_text = _load_rows(source)
    question_index = _find_header(headers, QUESTION_HEADERS)
    if question_index is None:
        raise ValueError("QA 文件缺少 Question/问题 列。")

    answer_index = _find_header(headers, ANSWER_HEADERS)
    number_index = _find_header(headers, NUMBER_HEADERS)
    explanation_index = _find_header(headers, EXPLANATION_HEADERS)
    category_index = _find_header(headers, CATEGORY_HEADERS)
    options_index = _find_header(headers, OPTION_HEADERS)
    unit_index = _find_header(headers, UNIT_HEADERS)

    known_indexes = {
        value
        for value in (
            question_index,
            answer_index,
            number_index,
            explanation_index,
            category_index,
            options_index,
            unit_index,
        )
        if value is not None
    }

    questions: list[QuestionRecord] = []
    for row_number, row in enumerate(rows, start=header_row + 1):
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        question = _stringify(padded[question_index])
        if not question:
            continue

        extra: dict[str, str] = {}
        for index, header in enumerate(headers):
            if index in known_indexes or index >= len(padded):
                continue
            value = _stringify(padded[index])
            if header and value:
                extra[header] = value

        location = f"{source.name}:row={row_number}"
        if sheet_name:
            location = f"{source.name}:sheet={sheet_name}:row={row_number}"

        questions.append(
            QuestionRecord(
                number=_stringify(padded[number_index]) if number_index is not None else str(len(questions) + 1),
                question=question,
                explanation=_stringify(padded[explanation_index]) if explanation_index is not None else "",
                category=_stringify(padded[category_index]) if category_index is not None else "",
                options=_split_options(_stringify(padded[options_index])) if options_index is not None else (),
                unit=_stringify(padded[unit_index]) if unit_index is not None else "",
                answer=_stringify(padded[answer_index]) if answer_index is not None else "",
                source_reference=location,
                row_number=row_number,
                extra=extra,
            )
        )

    if not questions:
        raise ValueError("QA 文件中没有识别到任何问题行。")
    return QuestionCatalog(
        source_path=str(source),
        sheet_name=sheet_name,
        header_row=header_row,
        questions=questions,
        preamble_text=preamble_text,
    )