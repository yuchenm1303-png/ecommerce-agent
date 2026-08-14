from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any, Iterable
from xml.etree import ElementTree

from openpyxl import load_workbook
from PIL import Image, UnidentifiedImageError

from .listing_images import select_listing_images
from .source_snapshot import SourceSnapshot, SnapshotTableRow, write_source_snapshot


PRODUCT_PACK_SCHEMA_VERSION = 1
PRODUCT_PACK_URL_HOST = "product-pack.invalid"
SUPPORTED_DOCUMENT_SUFFIXES = {
    ".pdf",
    ".docx",
    ".xlsx",
    ".xlsm",
    ".csv",
    ".tsv",
    ".txt",
    ".md",
}
SUPPORTED_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif"}
SUPPORTED_ARCHIVE_SUFFIXES = {".zip"}
SUPPORTED_PRODUCT_PACK_SUFFIXES = (
    SUPPORTED_DOCUMENT_SUFFIXES | SUPPORTED_IMAGE_SUFFIXES | SUPPORTED_ARCHIVE_SUFFIXES
)

_MAX_INPUT_FILES = 200
_MAX_SINGLE_FILE_BYTES = 64 * 1024 * 1024
_MAX_ARCHIVE_EXPANDED_BYTES = 512 * 1024 * 1024
_MAX_VISIBLE_TEXT_CHARS = 160_000
_MAX_TABLE_ROWS = 12_000


class ProductPackError(ValueError):
    pass


@dataclass(slots=True, frozen=True)
class ProductPackFile:
    original_path: str
    stored_path: str
    sha256: str
    size_bytes: int
    suffix: str
    kind: str

    def as_dict(self) -> dict[str, Any]:
        return {
            "original_path": self.original_path,
            "stored_path": self.stored_path,
            "sha256": self.sha256,
            "size_bytes": self.size_bytes,
            "suffix": self.suffix,
            "kind": self.kind,
        }


@dataclass(slots=True, frozen=True)
class ProductPackCapture:
    manifest_path: Path
    product_reference_url: str
    bootstrap_snapshot_path: Path
    bootstrap_snapshot: SourceSnapshot
    customer_snapshot_paths: tuple[Path, ...]
    evidence_image_paths: tuple[Path, ...]
    listing_image_paths: tuple[Path, ...]
    stored_files: tuple[ProductPackFile, ...]
    warnings: tuple[str, ...] = ()

    @property
    def snapshot_path(self) -> Path:
        return self.bootstrap_snapshot_path

    @property
    def snapshot(self) -> SourceSnapshot:
        return self.bootstrap_snapshot

    @property
    def product_image_paths(self) -> tuple[Path, ...]:
        return self.evidence_image_paths

    @property
    def cache_hit(self) -> bool:
        return False

    @property
    def launched_now(self) -> bool:
        return False


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_name(value: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z._()\-\u4e00-\u9fff]+", "_", value).strip("._")
    return cleaned[:120] or "file"


def _kind_for_suffix(suffix: str) -> str:
    lowered = suffix.casefold()
    if lowered in SUPPORTED_IMAGE_SUFFIXES:
        return "image"
    if lowered in SUPPORTED_DOCUMENT_SUFFIXES:
        return "document"
    if lowered in SUPPORTED_ARCHIVE_SUFFIXES:
        return "archive"
    return "unsupported"


def _validate_input_path(path: Path) -> None:
    if not path.is_file():
        raise ProductPackError(f"商品资料文件不存在：{path}")
    if path.suffix.casefold() not in SUPPORTED_PRODUCT_PACK_SUFFIXES:
        raise ProductPackError(f"不支持的商品资料格式：{path.name}")
    size = path.stat().st_size
    if size <= 0:
        raise ProductPackError(f"商品资料文件为空：{path.name}")
    if size > _MAX_SINGLE_FILE_BYTES:
        raise ProductPackError(
            f"单个商品资料文件超过 {_MAX_SINGLE_FILE_BYTES // (1024 * 1024)} MB：{path.name}"
        )


def _copy_input(path: Path, raw_dir: Path, ordinal: int) -> Path:
    raw_dir.mkdir(parents=True, exist_ok=True)
    target = raw_dir / f"{ordinal:03d}-{_safe_name(path.name)}"
    if target.exists():
        target.unlink()
    shutil.copy2(path, target)
    return target


def _safe_zip_member(name: str) -> PurePosixPath | None:
    path = PurePosixPath(name)
    if path.is_absolute() or ".." in path.parts:
        return None
    if not path.name:
        return None
    if path.suffix.casefold() not in (SUPPORTED_DOCUMENT_SUFFIXES | SUPPORTED_IMAGE_SUFFIXES):
        return None
    return path


def _expand_zip(path: Path, target_dir: Path) -> list[Path]:
    target_dir.mkdir(parents=True, exist_ok=True)
    extracted: list[Path] = []
    expanded_bytes = 0
    try:
        archive = zipfile.ZipFile(path)
    except zipfile.BadZipFile as exc:
        raise ProductPackError(f"ZIP 文件损坏：{path.name}") from exc
    with archive:
        for info in archive.infolist():
            safe = _safe_zip_member(info.filename)
            if safe is None or info.is_dir():
                continue
            if len(extracted) >= _MAX_INPUT_FILES:
                raise ProductPackError(f"ZIP 内支持的商品资料文件超过 {_MAX_INPUT_FILES} 个。")
            if info.file_size <= 0:
                continue
            if info.file_size > _MAX_SINGLE_FILE_BYTES:
                raise ProductPackError(f"ZIP 内文件过大：{info.filename}")
            expanded_bytes += int(info.file_size)
            if expanded_bytes > _MAX_ARCHIVE_EXPANDED_BYTES:
                raise ProductPackError("ZIP 解包后的支持文件总量超过安全上限。")
            target = target_dir / f"{len(extracted) + 1:03d}-{_safe_name(safe.name)}"
            with archive.open(info, "r") as source, target.open("wb") as output:
                shutil.copyfileobj(source, output, length=1024 * 1024)
            extracted.append(target)
    if not extracted:
        raise ProductPackError(f"ZIP 中没有可读取的商品文档/表格/图片：{path.name}")
    return extracted


def _decode_text(data: bytes, *, label: str) -> str:
    for encoding in ("utf-8-sig", "utf-16", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    try:
        return data.decode("latin-1")
    except UnicodeDecodeError as exc:
        raise ProductPackError(f"无法读取文本编码：{label}") from exc


def _snapshot_origin(path: Path, fragment: str = "") -> str:
    value = path.resolve().as_uri()
    return value + (f"#{fragment}" if fragment else "")


def _new_snapshot(
    *,
    path: Path,
    title: str,
    visible_text: str,
    table_rows: Iterable[SnapshotTableRow] = (),
    fragment: str = "",
    meta: dict[str, str] | None = None,
    warnings: Iterable[str] = (),
) -> SourceSnapshot:
    origin = _snapshot_origin(path, fragment)
    return SourceSnapshot(
        requested_url=origin,
        final_url=origin,
        title=title,
        captured_at=_utc_now(),
        visible_text=str(visible_text or "")[:_MAX_VISIBLE_TEXT_CHARS],
        table_rows=list(table_rows)[:_MAX_TABLE_ROWS],
        meta={
            "input_mode": "customer_product_pack",
            "file_name": path.name,
            **(meta or {}),
        },
        warnings=list(warnings),
    )


def _text_snapshot(path: Path) -> list[SourceSnapshot]:
    text = _decode_text(path.read_bytes(), label=path.name).strip()
    if not text:
        raise ProductPackError(f"文本文件没有可用内容：{path.name}")
    return [_new_snapshot(path=path, title=path.name, visible_text=text)]


def _csv_snapshot(path: Path, *, delimiter: str) -> list[SourceSnapshot]:
    text = _decode_text(path.read_bytes(), label=path.name)
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    nonempty = [row for row in rows if any(str(cell).strip() for cell in row)]
    if not nonempty:
        raise ProductPackError(f"表格文件没有可用内容：{path.name}")

    headers = [str(value).strip() for value in nonempty[0]]
    structured: list[SnapshotTableRow] = []
    rendered: list[str] = []
    for row_number, row in enumerate(nonempty, start=1):
        rendered.append(
            f"row {row_number}: "
            + " | ".join(str(value).strip() for value in row if str(value).strip())
        )
        if row_number == 1:
            continue
        for column_index, raw in enumerate(row, start=1):
            value = str(raw).strip()
            if not value:
                continue
            header = headers[column_index - 1] if column_index <= len(headers) else ""
            key = header or f"Column {column_index}"
            structured.append(
                SnapshotTableRow(
                    key=key,
                    value=value,
                    table_index=1,
                    row_index=row_number,
                )
            )
    return [
        _new_snapshot(
            path=path,
            title=path.name,
            visible_text="\n".join(rendered),
            table_rows=structured,
            meta={"table_format": path.suffix.casefold().lstrip(".")},
        )
    ]


def _excel_snapshots(path: Path) -> list[SourceSnapshot]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ProductPackError(f"无法读取 Excel：{path.name}: {exc}") from exc
    snapshots: list[SourceSnapshot] = []
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            rows = [
                tuple(cell for cell in row)
                for row in sheet.iter_rows(values_only=True)
                if any(cell not in (None, "") for cell in row)
            ]
            if not rows:
                continue
            headers = [str(value).strip() if value not in (None, "") else "" for value in rows[0]]
            structured: list[SnapshotTableRow] = []
            rendered: list[str] = [f"Sheet: {sheet.title}"]
            for data_index, row in enumerate(rows, start=1):
                values = ["" if value is None else str(value).strip() for value in row]
                rendered.append(
                    f"row {data_index}: " + " | ".join(value for value in values if value)
                )
                if data_index == 1:
                    continue
                nonempty = [(index, value) for index, value in enumerate(values, start=1) if value]
                if len(nonempty) == 2 and not all(headers):
                    structured.append(
                        SnapshotTableRow(
                            key=nonempty[0][1],
                            value=nonempty[1][1],
                            table_index=sheet_index,
                            row_index=data_index,
                        )
                    )
                    continue
                for column_index, value in nonempty:
                    header = headers[column_index - 1] if column_index <= len(headers) else ""
                    structured.append(
                        SnapshotTableRow(
                            key=header or f"Column {column_index}",
                            value=value,
                            table_index=sheet_index,
                            row_index=data_index,
                        )
                    )
            snapshots.append(
                _new_snapshot(
                    path=path,
                    title=f"{path.name} · {sheet.title}",
                    visible_text="\n".join(rendered),
                    table_rows=structured,
                    fragment=f"sheet={sheet_index}",
                    meta={"sheet_name": sheet.title, "sheet_index": str(sheet_index)},
                )
            )
    finally:
        workbook.close()
    if not snapshots:
        raise ProductPackError(f"Excel 没有可用工作表：{path.name}")
    return snapshots


def _docx_snapshots(path: Path, media_dir: Path) -> tuple[list[SourceSnapshot], list[Path]]:
    try:
        archive = zipfile.ZipFile(path)
    except zipfile.BadZipFile as exc:
        raise ProductPackError(f"Word 文件损坏：{path.name}") from exc

    paragraphs: list[str] = []
    rows: list[SnapshotTableRow] = []
    images: list[Path] = []
    with archive:
        try:
            document_xml = archive.read("word/document.xml")
        except KeyError as exc:
            raise ProductPackError(f"Word 缺少 document.xml：{path.name}") from exc
        root = ElementTree.fromstring(document_xml)
        namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
        for paragraph in root.iter(namespace + "p"):
            text = "".join(node.text or "" for node in paragraph.iter(namespace + "t")).strip()
            if text:
                paragraphs.append(text)
        table_index = 0
        for table in root.iter(namespace + "tbl"):
            table_index += 1
            row_index = 0
            for row in table.findall(namespace + "tr"):
                row_index += 1
                cells: list[str] = []
                for cell in row.findall(namespace + "tc"):
                    text = " ".join(
                        node.text or "" for node in cell.iter(namespace + "t") if (node.text or "").strip()
                    ).strip()
                    cells.append(text)
                nonempty = [value for value in cells if value]
                if len(nonempty) >= 2:
                    rows.append(
                        SnapshotTableRow(
                            key=nonempty[0],
                            value=" | ".join(nonempty[1:]),
                            table_index=table_index,
                            row_index=row_index,
                        )
                    )

        media_dir.mkdir(parents=True, exist_ok=True)
        for name in archive.namelist():
            posix = PurePosixPath(name)
            if len(posix.parts) < 3 or posix.parts[:2] != ("word", "media"):
                continue
            if posix.suffix.casefold() not in SUPPORTED_IMAGE_SUFFIXES:
                continue
            data = archive.read(name)
            if len(data) < 1024:
                continue
            digest = hashlib.sha256(data).hexdigest()
            target = media_dir / f"{_safe_name(path.stem)}-{len(images) + 1:03d}-{digest[:10]}{posix.suffix.casefold()}"
            target.write_bytes(data)
            images.append(target)

    if not paragraphs and not rows and not images:
        raise ProductPackError(f"Word 没有可读取内容：{path.name}")
    visible = "\n".join(paragraphs)
    snapshot = _new_snapshot(
        path=path,
        title=path.name,
        visible_text=visible or f"Word document: {path.name}",
        table_rows=rows,
        meta={"embedded_images": str(len(images))},
    )
    return [snapshot], images


def _pdf_snapshots(path: Path) -> list[SourceSnapshot]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - dependency contract
        raise ProductPackError("读取 PDF 需要 pypdf；请重新安装 requirements.txt。") from exc
    try:
        reader = PdfReader(str(path))
    except Exception as exc:
        raise ProductPackError(f"无法读取 PDF：{path.name}: {exc}") from exc
    snapshots: list[SourceSnapshot] = []
    empty_pages = 0
    for page_index, page in enumerate(reader.pages, start=1):
        try:
            text = str(page.extract_text() or "").strip()
        except Exception:
            text = ""
        if not text:
            empty_pages += 1
            continue
        snapshots.append(
            _new_snapshot(
                path=path,
                title=f"{path.name} · Page {page_index}",
                visible_text=text,
                fragment=f"page={page_index}",
                meta={"page": str(page_index), "page_count": str(len(reader.pages))},
            )
        )
    if not snapshots:
        raise ProductPackError(
            f"PDF 没有可提取文字：{path.name}。扫描版 PDF 请同时上传原始页面图片。"
        )
    if empty_pages:
        snapshots[0].warnings.append(
            f"{empty_pages} page(s) contained no extractable text; scanned pages require image evidence."
        )
    return snapshots


def _validate_image(path: Path) -> None:
    try:
        with Image.open(path) as image:
            image.verify()
    except (UnidentifiedImageError, OSError, ValueError) as exc:
        raise ProductPackError(f"无法读取图片：{path.name}") from exc


def _parse_document(path: Path, media_dir: Path) -> tuple[list[SourceSnapshot], list[Path]]:
    suffix = path.suffix.casefold()
    if suffix in {".txt", ".md"}:
        return _text_snapshot(path), []
    if suffix == ".csv":
        return _csv_snapshot(path, delimiter=","), []
    if suffix == ".tsv":
        return _csv_snapshot(path, delimiter="\t"), []
    if suffix in {".xlsx", ".xlsm"}:
        return _excel_snapshots(path), []
    if suffix == ".docx":
        return _docx_snapshots(path, media_dir)
    if suffix == ".pdf":
        return _pdf_snapshots(path), []
    raise ProductPackError(f"不支持的商品文档：{path.name}")


def _write_snapshots(snapshots: Iterable[SourceSnapshot], snapshot_dir: Path) -> tuple[Path, ...]:
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    output: list[Path] = []
    for index, snapshot in enumerate(snapshots, start=1):
        path = snapshot_dir / f"customer-source-{index:04d}.json"
        output.append(write_source_snapshot(snapshot, path))
    return tuple(output)


def _bootstrap_snapshot(snapshots: list[SourceSnapshot], files: list[ProductPackFile]) -> SourceSnapshot:
    text_parts: list[str] = []
    rows: list[SnapshotTableRow] = []
    used = 0
    for snapshot in snapshots:
        if snapshot.visible_text.strip() and used < _MAX_VISIBLE_TEXT_CHARS:
            block = f"[{snapshot.title}]\n{snapshot.visible_text.strip()}"
            remaining = _MAX_VISIBLE_TEXT_CHARS - used
            text_parts.append(block[:remaining])
            used += min(len(block), remaining)
        if len(rows) < _MAX_TABLE_ROWS:
            rows.extend(snapshot.table_rows[: _MAX_TABLE_ROWS - len(rows)])
    if not text_parts:
        text_parts.append(
            "Customer product pack files:\n" + "\n".join(file.stored_path for file in files)
        )
    return SourceSnapshot(
        requested_url="product-pack://local",
        final_url="product-pack://local",
        title="Customer Product Pack",
        captured_at=_utc_now(),
        visible_text="\n\n".join(text_parts),
        table_rows=rows,
        meta={
            "input_mode": "customer_product_pack",
            "stored_files": str(len(files)),
            "document_snapshots": str(len(snapshots)),
        },
    )


def _pack_reference(files: Iterable[ProductPackFile]) -> str:
    digest = hashlib.sha256()
    for item in sorted(files, key=lambda value: (value.sha256, value.stored_path)):
        digest.update(item.sha256.encode("ascii"))
        digest.update(b"\0")
        digest.update(item.stored_path.encode("utf-8"))
        digest.update(b"\0")
    return f"https://{PRODUCT_PACK_URL_HOST}/{digest.hexdigest()[:24]}"


def capture_product_pack(
    paths: Iterable[str | Path],
    *,
    output_dir: str | Path,
) -> ProductPackCapture:
    """Persist and normalize one customer-supplied product evidence pack.

    Documents are converted into citable SourceSnapshot units, images remain exact
    byte evidence, and the original bytes are copied under the run directory. No
    marketplace-field semantics are inferred here; this is mechanical intake only.
    """

    originals = [Path(value).expanduser() for value in paths]
    if not originals:
        raise ProductPackError("请至少选择一个商品资料文件。")
    if len(originals) > _MAX_INPUT_FILES:
        raise ProductPackError(f"一次最多接收 {_MAX_INPUT_FILES} 个商品资料文件。")
    for path in originals:
        _validate_input_path(path)

    root = Path(output_dir).resolve()
    if root.exists():
        shutil.rmtree(root)
    root.mkdir(parents=True, exist_ok=True)
    raw_dir = root / "raw"
    unpacked_dir = root / "unpacked"
    snapshot_dir = root / "snapshots"
    media_dir = root / "embedded-images"

    stored_files: list[ProductPackFile] = []
    parse_paths: list[Path] = []
    original_by_stored: dict[Path, str] = {}
    seen_sha: set[str] = set()

    for ordinal, original in enumerate(originals, start=1):
        stored = _copy_input(original, raw_dir, ordinal)
        digest = _sha256_path(stored)
        if digest in seen_sha:
            stored.unlink(missing_ok=True)
            continue
        seen_sha.add(digest)
        stored_files.append(
            ProductPackFile(
                original_path=str(original.resolve()),
                stored_path=str(stored.resolve()),
                sha256=digest,
                size_bytes=stored.stat().st_size,
                suffix=stored.suffix.casefold(),
                kind=_kind_for_suffix(stored.suffix),
            )
        )
        if stored.suffix.casefold() == ".zip":
            expanded = _expand_zip(stored, unpacked_dir / f"archive-{ordinal:03d}")
            for child in expanded:
                child_digest = _sha256_path(child)
                if child_digest in seen_sha:
                    continue
                seen_sha.add(child_digest)
                stored_files.append(
                    ProductPackFile(
                        original_path=f"{original.resolve()}::{child.name}",
                        stored_path=str(child.resolve()),
                        sha256=child_digest,
                        size_bytes=child.stat().st_size,
                        suffix=child.suffix.casefold(),
                        kind=_kind_for_suffix(child.suffix),
                    )
                )
                parse_paths.append(child)
                original_by_stored[child] = str(original.resolve())
        else:
            parse_paths.append(stored)
            original_by_stored[stored] = str(original.resolve())

    snapshots: list[SourceSnapshot] = []
    images: list[Path] = []
    warnings: list[str] = []
    for path in parse_paths:
        if path.suffix.casefold() in SUPPORTED_IMAGE_SUFFIXES:
            _validate_image(path)
            images.append(path)
            continue
        try:
            parsed, embedded_images = _parse_document(path, media_dir)
        except ProductPackError as exc:
            warnings.append(str(exc))
            continue
        snapshots.extend(parsed)
        for image in embedded_images:
            _validate_image(image)
            images.append(image)

    if not snapshots and not images:
        details = " | ".join(warnings) if warnings else "没有可解析内容"
        raise ProductPackError("商品资料包没有形成可用文本/表格/图片证据：" + details)

    snapshot_paths = _write_snapshots(snapshots, snapshot_dir)
    bootstrap = _bootstrap_snapshot(snapshots, stored_files)
    bootstrap_path = write_source_snapshot(bootstrap, root / "bootstrap-source.json")

    image_dedup: dict[str, Path] = {}
    for image in images:
        image_dedup.setdefault(_sha256_path(image), image)
    evidence_images = tuple(image_dedup.values())
    listing_selection = select_listing_images(evidence_images)
    listing_images = tuple(listing_selection.selected)

    reference_url = _pack_reference(stored_files)
    manifest = {
        "schema_version": PRODUCT_PACK_SCHEMA_VERSION,
        "input_mode": "customer_product_pack",
        "product_reference_url": reference_url,
        "created_at": _utc_now(),
        "stored_files": [item.as_dict() for item in stored_files],
        "bootstrap_snapshot": str(bootstrap_path.resolve()),
        "customer_snapshots": [str(path.resolve()) for path in snapshot_paths],
        "evidence_images": [str(path.resolve()) for path in evidence_images],
        "listing_images": [str(path.resolve()) for path in listing_images],
        "listing_image_rejected": listing_selection.rejected_count,
        "warnings": warnings,
    }
    manifest_path = root / "product-pack.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    return ProductPackCapture(
        manifest_path=manifest_path,
        product_reference_url=reference_url,
        bootstrap_snapshot_path=bootstrap_path,
        bootstrap_snapshot=bootstrap,
        customer_snapshot_paths=snapshot_paths,
        evidence_image_paths=evidence_images,
        listing_image_paths=listing_images,
        stored_files=tuple(stored_files),
        warnings=tuple(warnings),
    )


def load_product_pack_manifest(path: str | Path) -> dict[str, Any]:
    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(f"商品资料包 manifest 不存在：{source}")
    payload = json.loads(source.read_text(encoding="utf-8"))
    if not isinstance(payload, dict) or payload.get("schema_version") != PRODUCT_PACK_SCHEMA_VERSION:
        raise ProductPackError("商品资料包 manifest 版本无效。")
    if payload.get("input_mode") != "customer_product_pack":
        raise ProductPackError("商品资料包 manifest input_mode 无效。")
    reference = str(payload.get("product_reference_url") or "").strip()
    if not reference.startswith(f"https://{PRODUCT_PACK_URL_HOST}/"):
        raise ProductPackError("商品资料包 product_reference_url 无效。")
    for key in ("bootstrap_snapshot",):
        value = Path(str(payload.get(key) or ""))
        if not value.is_file():
            raise ProductPackError(f"商品资料包缺少 {key}：{value}")
    for key in ("customer_snapshots", "evidence_images", "listing_images"):
        values = payload.get(key) or []
        if not isinstance(values, list):
            raise ProductPackError(f"商品资料包 {key} 必须是数组。")
        missing = [str(value) for value in values if not Path(str(value)).is_file()]
        if missing:
            raise ProductPackError(f"商品资料包 {key} 文件缺失：" + " | ".join(missing[:5]))
    return payload


__all__ = [
    "PRODUCT_PACK_SCHEMA_VERSION",
    "ProductPackCapture",
    "ProductPackError",
    "ProductPackFile",
    "SUPPORTED_PRODUCT_PACK_SUFFIXES",
    "capture_product_pack",
    "load_product_pack_manifest",
]
