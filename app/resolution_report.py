from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .resolution_engine import ResolutionRecord, summarize_resolution


REPORT_COLUMNS = (
    "No.",
    "Question",
    "Category",
    "Explanation",
    "Options",
    "Unit",
    "Answer",
    "Status",
    "Auto Fill",
    "Confidence",
    "Source Type",
    "Source Reference",
    "Evidence",
    "Detail",
    "Provenance JSON",
)


def _row(record: ResolutionRecord) -> list[object]:
    return [
        record.question_number,
        record.label,
        record.question_category,
        record.question_explanation,
        " | ".join(record.question_options),
        record.question_unit,
        record.answer or "",
        record.status,
        "YES" if record.eligible_for_autofill else "NO",
        round(record.confidence, 4),
        record.source_type or "",
        record.source_reference or "",
        record.evidence or "",
        record.detail,
        json.dumps(record.provenance, ensure_ascii=False),
    ]


def write_resolution_json(records: Iterable[ResolutionRecord], path: str | Path) -> Path:
    target = Path(path)
    items = list(records)
    payload = {
        "summary": summarize_resolution(items),
        "records": [item.as_dict() for item in items],
    }
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return target


def write_resolution_xlsx(records: Iterable[ResolutionRecord], path: str | Path) -> Path:
    target = Path(path)
    items = list(records)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resolution"
    sheet.append(list(REPORT_COLUMNS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(REPORT_COLUMNS))}{max(1, len(items) + 1)}"

    for item in items:
        sheet.append(_row(item))

    widths = {
        1: 8,
        2: 32,
        3: 22,
        4: 42,
        5: 38,
        6: 14,
        7: 28,
        8: 16,
        9: 12,
        10: 12,
        11: 20,
        12: 52,
        13: 34,
        14: 54,
        15: 70,
    }
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width

    summary = workbook.create_sheet("Summary")
    stats = summarize_resolution(items)
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    for key, value in stats.items():
        summary.append([key, value])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 16

    workbook.save(target)
    return target
