from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .models import ProductRecord


SKU_HEADER_ALIASES = {"sku", "商品sku", "商品编码", "商品编号", "itemsku"}


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: object) -> str:
    return _stringify(value).lower().replace(" ", "").replace("_", "")


def _find_sku_header(headers: Iterable[str]) -> str:
    for header in headers:
        if _normalize_header(header) in SKU_HEADER_ALIASES:
            return header
    raise ValueError(
        "未找到 SKU 列。请至少包含一个表头：SKU / 商品SKU / 商品编码 / 商品编号。"
    )


def _records_from_rows(headers: list[str], rows: Iterable[list[object]]) -> list[ProductRecord]:
    sku_header = _find_sku_header(headers)
    records: list[ProductRecord] = []

    for row_number, row in enumerate(rows, start=2):
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        values = {header: _stringify(value) for header, value in zip(headers, padded)}
        sku = values.get(sku_header, "").strip()
        if not sku:
            continue

        product_values = {
            header: value
            for header, value in values.items()
            if header != sku_header and value != ""
        }
        records.append(ProductRecord(sku=sku, values=product_values, row_number=row_number))

    if not records:
        raise ValueError("表格中没有可执行的商品数据。")
    return records


def _load_csv(path: Path) -> list[ProductRecord]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            raw_headers = next(reader)
        except StopIteration as exc:
            raise ValueError("CSV 文件为空。") from exc

        headers = [_stringify(item) for item in raw_headers]
        if any(not item for item in headers):
            raise ValueError("CSV 表头不能为空。")
        return _records_from_rows(headers, reader)


def _load_xlsx(path: Path) -> list[ProductRecord]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise ValueError("Excel 文件为空。") from exc

        headers = [_stringify(item) for item in raw_headers]
        if any(not item for item in headers):
            raise ValueError("Excel 表头不能为空。")
        return _records_from_rows(headers, (list(row) for row in rows))
    finally:
        workbook.close()


def load_products(path: str | Path) -> list[ProductRecord]:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"找不到数据文件：{source}")

    suffix = source.suffix.lower()
    if suffix == ".csv":
        return _load_csv(source)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx(source)

    raise ValueError("当前仅支持 .csv / .xlsx / .xlsm 表格。")
